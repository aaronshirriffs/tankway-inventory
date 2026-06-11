"""
Scheduled email export engine — delivers the same filtered data as
/v1/products to customers who can't consume an API, as an emailed XLSX/CSV.

Data comes from the unmodified build_products() (imported late from app to
avoid a circular import), so filters, stock, lead times, incoming, and prices
are byte-for-byte the API's own logic. Export-only columns (weight,
categories) come from odoo_client.fetch_export_extras(), which the API path
never calls.

Per-key settings live in keys.json under "export" (same pattern as
show_price / show_incoming); absent or disabled => nothing is ever sent.
"""
import csv
import io
import logging
import re
from datetime import datetime
from zoneinfo import ZoneInfo

import columns
import gmail_client
import odoo_client
import storage

logger = logging.getLogger(__name__)

NZT = ZoneInfo("Pacific/Auckland")

DEFAULT_EXPORT = {
    "enabled": False,
    "format": "xlsx",            # 'xlsx' | 'csv' — fixed per customer
    "email": "",                 # recipient(s), comma-separated
    "reply_to": "",              # per-customer Reply-To (sender is always the connected account)
    "frequency": "weekly",       # 'daily' | 'weekly'
    "weekday": 0,                # 0=Monday .. 6=Sunday (weekly only)
    "hour": 7,                   # NZT hour 0-23
    "columns": {t: False for t in columns.COLUMN_TOGGLES},
    "last_sent": None,           # ISO UTC of the last successful send
    "last_result": "",
}


class ExportError(Exception):
    """Configuration problem the admin can fix (shown verbatim in the UI)."""


def export_config(config):
    """Key's export settings merged over defaults (missing block => disabled)."""
    merged = {k: (dict(v) if isinstance(v, dict) else v) for k, v in DEFAULT_EXPORT.items()}
    ex = config.get("export")
    if isinstance(ex, dict):
        for k, v in ex.items():
            if k == "columns":
                if isinstance(v, dict):
                    merged["columns"].update(v)
                # non-dict columns value: ignore, keep safe defaults
            else:
                merged[k] = v
    return merged


def _slug(label):
    return re.sub(r"[^A-Za-z0-9]+", "_", label or "customer").strip("_") or "customer"


def _render_xlsx(headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    for idx, h in enumerate(headers, start=1):
        width = max([len(str(h))] + [len(str(r[idx - 1])) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 60)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_csv(headers, rows):
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow(headers)
    w.writerows(rows)
    # UTF-8 BOM so Excel opens it with correct encoding by default.
    return buf.getvalue().encode("utf-8-sig")


def build_export_file(config):
    """-> (filename, data_bytes, fmt, product_count). No email is sent."""
    from app import build_products  # late import; app is fully loaded by call time

    ecfg = export_config(config)
    on = ecfg["columns"]

    # Work on a copy so the engine returns the fields the selected columns
    # need. The export is driven ENTIRELY by its own column toggles — the key's
    # API-facing visibility toggles (Show sales price / Show incoming stock)
    # have no effect on the export. The stored config — and therefore the live
    # API — is never modified.
    working = dict(config)
    working["show_price"] = bool(on.get("price_exc") or on.get("price_inc"))
    working["show_incoming"] = bool(on.get("incoming"))
    products = build_products(working)

    # extras are always fetched now: the Name column appends each product's
    # variant suffix (so variants are distinguishable), and the email body
    # reports how many distinct inventory categories the file spans.
    extras = {}
    if products:
        uid, models = odoo_client.connect()
        extras = odoo_client.fetch_export_extras(models, uid, [p["id"] for p in products])

    headers, rows = columns.build_rows(products, config, extras, on)
    ncat = len({e.get("inv_category") for e in extras.values() if e.get("inv_category")})
    fmt = "csv" if ecfg["format"] == "csv" else "xlsx"
    data = _render_csv(headers, rows) if fmt == "csv" else _render_xlsx(headers, rows)
    stamp = datetime.now(NZT).strftime("%Y-%m-%d")
    filename = f"MDR_Stock_{_slug(config.get('label'))}_{stamp}.{fmt}"
    return filename, data, fmt, len(products), ncat


def run_export(token, config):
    """Build and email one key's export. Returns {to, products, filename}."""
    ecfg = export_config(config)
    to = (ecfg.get("email") or "").strip()
    if not to:
        raise ExportError("No recipient email is configured for this key.")
    if not gmail_client.is_authorised():
        raise ExportError("Gmail is not connected yet — run auth_setup_export.py once.")

    filename, data, fmt, count, ncat = build_export_file(config)
    stamp = datetime.now(NZT).strftime("%d %B %Y")
    cust = (config.get("label") or "").strip() or "there"
    cat_word = "category" if ncat == 1 else "categories"
    body = (
        f"Hi {cust},\n\n"
        f"Please find attached the current stock availability information from MDR Lighting "
        f"({count} products, {ncat} {cat_word}, generated {stamp}).\n\n"
        f"Regards,\nMDR Lighting\n"
    )
    gmail_client.send_with_attachment(
        to=to,
        subject=f"MDR Lighting stock update — {stamp}",
        body=body,
        filename=filename,
        data=data,
        fmt=fmt,
        reply_to=(ecfg.get("reply_to") or "").strip() or None,
    )
    _record_result(token, sent=True, note=f"sent {count} products to {to}")
    return {"to": to, "products": count, "filename": filename}


def _record_result(token, sent, note):
    """Stamp last_sent/last_result on the key's CURRENT export block."""
    cfg = storage.get_key(token)
    if not cfg:
        return
    ecfg = export_config(cfg)
    if sent:
        ecfg["last_sent"] = storage.now_iso()
    ecfg["last_result"] = f"{storage.now_iso()} {note}"[:300]
    storage.update_key(token, export=ecfg)


def _is_due(ecfg, now):
    if not ecfg.get("enabled") or not (ecfg.get("email") or "").strip():
        return False
    if ecfg["frequency"] == "weekly" and now.weekday() != int(ecfg.get("weekday") or 0):
        return False
    if now.hour < int(ecfg.get("hour") or 0):
        return False
    last = ecfg.get("last_sent")
    if last:
        try:
            last_nz = datetime.fromisoformat(last).astimezone(NZT)
            if last_nz.date() == now.date():
                return False  # already sent today
        except ValueError:
            pass
    return True


def run_due_exports():
    """Scheduler entry point — never raises (live API must stay unaffected)."""
    try:
        keys = storage.load_keys()
    except Exception:
        logger.exception("export scheduler: could not load keys")
        return
    now = datetime.now(NZT)
    for token, config in keys.items():
        try:
            if not config.get("enabled"):
                continue  # disabled API keys never export either
            if _is_due(export_config(config), now):
                result = run_export(token, config)
                logger.info("export sent for %s: %s", config.get("label"), result)
        except Exception as e:
            logger.exception("export failed for %s", config.get("label"))
            try:
                _record_result(token, sent=False, note=f"FAILED: {e}")
            except Exception:
                pass
